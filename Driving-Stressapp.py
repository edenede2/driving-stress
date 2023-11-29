import streamlit as st
import pandas as pd
import re
import base64
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import PatternFill, Font, Color, Alignment
from openpyxl.comments import Comment
import plotly.express as px

HIGHLIGHT_VALUES = {
    'A': [1592, 2923, 3082, 3500, 3940, 4705, 5053, 4430, 6580],
    'B': [1032, 1980, 2661, 3250, 4332, 5560, 5845, 5945, 6487, 7850],
    'C': [670, 1300, 2513, 3457, 4107, 4390, 5037, 5358, 6484]
}
def process_raw_file_for_streamlit(uploaded_file, original_file_name):
    file_content = uploaded_file
    structured_data = extract_structured_data_v6(file_content)
    df_sorted = construct_dataframe_optimized_v2_refined(file_content, structured_data, original_file_name)
    return df_sorted

def save_as_xlsx_with_highlight_refined(df, scenario, file_name):
    """
    Save the DataFrame as an XLSX file and highlight rows based on Distm values and scenario.
    Also, updates the Event column based on highlighted rows.
    """
    # Define the fill pattern for highlighting
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    numeric_columns = ['Time', 'Velm', 'Distm', 'Xm', 'IncVm', 'IncRm', 'WheeleAng', 'ThrAcce', 'BrakAcce', 'TL', 'Crashes']
    for col in numeric_columns:
        df[col] = pd.to_numeric(df[col], errors='coerce', downcast='float')

    new_filename = f"sorted_{file_name.split('.')[0]}.xlsx"
    with pd.ExcelWriter(new_filename, engine='openpyxl') as writer:  
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        # Get the workbook and sheet for further editing
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Define the fill pattern for specific cells
        red_cells = PatternFill(start_color="00FF8080", end_color="00FF8080", fill_type="solid")  # Red color
        green_cells = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        blue_cells = PatternFill(start_color="0000CCFF", end_color="0000CCFF", fill_type="solid")
        Purple_cells = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Apply formatting to specific header cells
        red_headers = ['Time', 'Velm']  # List of headers to color, adjust as needed
        green_headers = ['Eve2Y', 'Eve2X', 'Eve2Vel', 'Eve4Y', 'Eve4X', 'Eve4Vel', 'Eve4Y', 'Eve6X', 'Eve6Vel', 'Eve8Y', 'Eve8X', 'Eve8Vel']
        blue_headers = ['IncVm', 'IncRm', 'WheeleAng', 'ThrAcce', 'BrakAcce','TL','Crashes','???', 'VelKPH','CompTime','YawRate','WheeleOP','ThrOP','BrakOP','WheeleYawRate','InertialAng','OPState']
        purple_headers = ['FirstRT', 'First Ditance']
  
        # Formatting header row
        for cell in worksheet[1]:  # worksheet[1] is the first row (header row)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.font = Font(bold=False)
            cell.alignment = Alignment(horizontal='center')

            # Add specific labels to certain cells if needed
            # Add specific labels to certain cells if needed
            if cell.value == "Crashes":
                comment = Comment("1 = Vehicle collisions; 2 = Off road collisions; 3 = Collisions with pedestrians; 4 = Collisions with lane markers (barrels, cones, etc.); 5 = Collisions with Jersey Barriers; 6 = Collisions with collision blocks", "Eden")
                cell.comment = comment
            elif cell.value == "TL":
                comment = Comment("0 = None, 1 = green; 2 = orange; 3 = red", "Eden")
                cell.comment = comment
            elif cell.value == "Participant":
                comment = Comment("Participant number (e.g., sub_001 = 001)", "Eden")
                cell.comment = comment
            elif cell.value == "Scenario":
                comment = Comment("The type of the driving scenario (e.g., A/B/C)", "Eden")
                cell.comment = comment
            elif cell.value == "Order":
                comment = Comment("Number of the scenario in order from 1 to 3", "Eden")
                cell.comment = comment
            elif cell.value == "Event":
                comment = Comment("The number of the event per scenario", "Eden")
                cell.comment = comment
            elif cell.value == "Time":
                comment = Comment("Time from start in seconds", "Eden")
                cell.comment = comment
            elif cell.value == "Velm":
                comment = Comment("Velocity (distm/Time^2), represents speed", "Eden")
                cell.comment = comment
            elif cell.value == "Distm":
                comment = Comment("Distance from start", "Eden")
                cell.comment = comment
            elif cell.value == "Xm":
                comment = Comment("Horizontal location (0 = middle of the road)", "Eden")
                cell.comment = comment
            elif cell.value == "WheeleAng":
                comment = Comment("The movement of the wheel", "Eden")
                cell.comment = comment
            elif cell.value == "ThrAcce":
                comment = Comment("Gas pedal", "Eden")
                cell.comment = comment
            elif cell.value == "BrakAcce":
                comment = Comment("Brake pedal", "Eden")
                cell.comment = comment
            elif cell.value == "FirstRT":
                comment = Comment("First reaction time (event time - reaction time)", "Eden")
                cell.comment = comment        
            elif cell.value == "First Ditance":
                comment = Comment("First reaction distance (event distance - reaction distance)", "Eden")
                cell.comment = comment



        
        for col_num, column_title in enumerate(worksheet[1], start=1):  # worksheet[1] is the header row
            cell = worksheet.cell(row=1, column=col_num)
            if column_title.value in red_headers:
                cell.fill = red_cells
            elif column_title.value in green_headers:
                cell.fill = green_cells
            elif column_title.value in blue_headers:
                cell.fill = blue_cells
            elif column_title.value in purple_headers:
                cell.fill = Purple_cells


        
        # Iterate over the rows to highlight rows with an 'Event'
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=1):
            event_value = worksheet.cell(row=row_idx, column=4).value
            if event_value:  # If there's an event value, highlight the row
                for cell in row:
                    cell.fill = highlight_fill
                    
    return new_filename

# "save_as_xlsx_with_highlight_refined"



# Extract structured data from raw file
def extract_structured_data_v6(file_content):
    structured_data = []
    start_reading = False
    skip_count = 2  # Number of lines to skip after the "Block #1: output_data," line
    short_line_count = 0  # Counter to track consecutive short lines
    
    for line in file_content.split('\n'):
        # If the line contains "Block #1: output_data,", prepare to start reading the structured data
        if "Block #1:   output_data," in line:
            start_reading = True
            continue  # skip the current line
        
        # If we've started reading
        if start_reading:
            # Skip the next two lines after "Block #1: output_data,"
            if skip_count > 0:
                skip_count -= 2
                continue

            # If the line has a reasonable length (indicative of structured data)
            if len(line) > 100:
                structured_data.append(line.strip())
                short_line_count = 0  # Reset the short line counter
            else:
                short_line_count += 1  # Increment the short line counter
            
                # If we encounter two consecutive short lines, stop the extraction
                if short_line_count >= 2:
                    break

    return structured_data

# Determine the header for the sorted data based on the 5th row of the txt file
def determine_header(file_content):
    lines = file_content.split('\n')
    fifth_line = lines[4].strip()

    if "Scenario1" in fifth_line:
        return pd.read_csv("TestA.csv")
    elif "Senario2" in fifth_line:
        return pd.read_csv("TestB.csv")
    elif "Scenario3" in fifth_line:
        return pd.read_csv("TestC.csv")
    else:
        raise ValueError("Unrecognized scenario in file.")

# Construct and populate the dataframe
def construct_dataframe_optimized_v2_refined(file_content, structured_data, original_file_name):
    header_df = determine_header(file_content)
    
    participant_number, order = original_file_name.replace(".txt", "").split('_')
    participant_number = str(participant_number)
    order = int(order)
    
    lines = file_content.split('\n')
    fifth_line = lines[4].strip()
   
    if "Scenario1" in fifth_line:
        scenario = "A"
    elif "Senario2" in fifth_line:
        scenario = "B"
    elif "Scenario3" in fifth_line:
        scenario = "C"
    else:
        scenario = None
    
    rows = []
    for data_row in structured_data:
        values = data_row.split()
        row_data = {
            'Participant': participant_number,
            'Scenario': scenario,
            'Order': order,
            'Event': None,
            'Time': values[0],
            'Velm': values[1],
            'Distm': float(values[2]),
            'Xm': values[3],
            'IncVm': values[4],
            'IncRm': values[5],
            'WheeleAng': values[6],
            'ThrAcce': values[7],
            'BrakAcce': values[8],
            'TL': values[9],
            'Crashes': values[10],
            'First RT': None,
            'First distance': None
        }
        for i, col in enumerate(header_df.columns[17:], start=11):
            if i < len(values):
                row_data[col] = values[i]
            else:
                row_data[col] = None
        rows.append(row_data)
    
    df = pd.DataFrame(rows, columns=header_df.columns)
    df['Distm'] = pd.to_numeric(df['Distm'], errors='coerce')
    for highlight_value in HIGHLIGHT_VALUES.get(scenario, []):
        closest_row_idx = (df['Distm'] - highlight_value).abs().idxmin()
        df.at[closest_row_idx, 'Event'] = df.at[closest_row_idx, 'Distm']
    
    df['Event'] = df['Event'].rank(method='first').fillna(0).astype(int)
    df['Event'] = df['Event'].replace(0, np.nan)
    
    return df


def get_event_description(scenario, event_number):
    """
    Get the description of the event based on the scenario and event number.
    """
    descriptions = {
        'A': {
            1: "1. changed traffic light.",
            2: "2. changed traffic light 2.",
            3: "3. car integrate from the right side.",
            4: "4. changed traffic light 3.",
            5: "5. A pedestrian crosses from the right - without a traffic light.",
            6: "6. A motorcyclist enters from the right.",
            7: "7. Traffic jam after a motorcycle and a stuck car.",
            8: "8. A pedestrian or animal crosses a road from the right and then a yellow vehicle drives slowly in the left lane and when the driver tries to overtake on the right he suddenly turns to the right lane and then goes down to the right to the curb.",
            9: "9. A car drives backwards and enters from the left"
        },
        'B': {
            1: "1. Pedestrian or animal crossing a road from the left.",
            2: "2. A motorcyclist enters from the left",
            3: "3. A light blue vehicle exits from the right and enters both lanes and then returns",
            4: "4. A vehicle from the separator - enters the lane (next to trucks in an industrial area)",
            5: "5. A traffic light turns red - at the end of the bridge",
            6: "6. A traffic light turns red - at the beginning of the city",
            7: "7. A cyclist disobeys a red light and enters the intersection on the right",
            8: "8. A pedestrian crosses from the right at a crosswalk without a traffic light",
            9: "9. Traffic light turns red",
            10: "10. At the end of the gas station and exit the scenario to the right"
        },
        'C': {
            1: "1. A traffic light turns red just before the bridge that enters the city",
            2: "2. A garbage truck from which garbage fell is blocking the left lane. A police car is behind and bursts into the right lane as the driver approaches",
            3: "3. A pedestrian crosses from the left at a crosswalk without a traffic light",
            4: "4. A bicycle enters the lane on the right",
            5: "5. A work vehicle leaves the parking lot and enters both lanes",
            6: "6. A vehicle parked on the right enters and crosses both lanes",
            7: "7. A traffic light turns red in the middle of the city",
            8: "8. A pedestrian in a crosswalk enters from the right",
            9: "9. Last traffic light in town - turns red"
        }
    }
    return descriptions.get(scenario, {}).get(event_number, "Unknown Event")


def calculate_changes(df, event_row_index, offset):
    """
    Calculate the changes in WheeleAng, ThrAcce, and BrakAcce for the selected row relative to the event row.
    """
    df['WheeleAng'] = pd.to_numeric(df['WheeleAng'], errors='coerce')
    df['ThrAcce'] = pd.to_numeric(df['ThrAcce'], errors='coerce')
    df['BrakAcce'] = pd.to_numeric(df['BrakAcce'], errors='coerce')
    df['Time'] = pd.to_numeric(df['Time'], errors='coerce')
    df['TL'] = pd.to_numeric(df['TL'], errors='coerce')
    event_row = df.iloc[event_row_index]
    target_row = df.iloc[event_row_index + offset]
    
    # Ensure the values are numeric before performing the subtraction
    def safe_subtract(val1, val2):
        try:
            return float(val1) - float(val2)
        except ValueError:
            return None  # or 0, or however you want to handle this case
    
    changes = {
        'WheeleAng': safe_subtract(target_row['WheeleAng'], event_row['WheeleAng']),
        'ThrAcce': safe_subtract(target_row['ThrAcce'], event_row['ThrAcce']),
        'BrakAcce': safe_subtract(target_row['BrakAcce'], event_row['BrakAcce']),
        'TimeDifference': safe_subtract(target_row['Time'], event_row['Time']),
        'DistmDifference': safe_subtract(target_row['Distm'], event_row['Distm'])
    }
    return changes


def plot_event_analysis_updated(df, selected_event, parameter, offset, pre_event_range, post_event_range):
    """
    Plot the change in the selected parameter around the event using Plotly with both line and scatter plots.
    Custom hover information is added to show distm, delta distm, time, and delta time.
    """
    df[parameter] = pd.to_numeric(df[parameter], errors='coerce')
    df['Time'] = pd.to_numeric(df['Time'], errors='coerce')
    df['Distm'] = pd.to_numeric(df['Distm'], errors='coerce')
    df['TL'] = pd.to_numeric(df['TL'], errors='coerce')  # Ensure TL is numeric

    # Find the index of the selected event
    event_index = df[df['Event'] == selected_event].index[0]
    event_distm = df.loc[event_index, 'Distm']
    event_time = df.loc[event_index, 'Time']

    # Extract a window of rows around the event
    window_start = max(0, event_index - pre_event_range)
    window_end = min(df.shape[0], event_index + post_event_range)
    df_window = df.iloc[window_start:window_end]

    # Calculate changes for 'BrakAcce' and custom hover text
    if parameter == 'BrakAcce':
        df_window['BrakAcce_change'] = df_window['BrakAcce'].diff().abs()
        marker_colors = np.where(df_window['BrakAcce_change'] >= 1.4, 'red', 'blue')
    else:
        marker_colors = 'blue'  # Default color

    hover_text = []
    for idx, row in df_window.iterrows():
        delta_distm = event_distm - row['Distm']
        delta_time = event_time - row['Time']
        hover_info = f'Distm: {row["Distm"]}<br>Delta Distm: {delta_distm}<br>Time: {row["Time"]}<br>Delta Time: {delta_time}'
        hover_text.append(hover_info)

    # Create Plotly line graph
    fig = px.line(df_window, x='Distm', y=parameter, hover_name=hover_text)

    # Add scatter plot with conditional coloring and custom hover text
    fig.add_scatter(x=df_window['Distm'], y=df_window[parameter], mode='markers', marker=dict(color=marker_colors), hovertext=hover_text)

    # Highlight the event line
    fig.add_vline(x=event_distm, line_dash="dash", line_color="red")

    # Traffic light color mapping and vertical lines for changes
    tl_color_map = {1: 'green', 2: 'orange', 3: 'red'}
    prev_tl = None
    for idx, row in df_window.iterrows():
        if row['TL'] != prev_tl:
            tl_color = tl_color_map.get(row['TL'], 'grey')
            fig.add_vline(x=row['Distm'], line_dash="dot", line_color=tl_color)
            prev_tl = row['TL']

    # Set titles and labels
    fig.update_layout(title=f'Change in {parameter} around Event {selected_event}', xaxis_title='Distm (Distance)', yaxis_title=parameter)

    # Display the plot in Streamlit
    st.plotly_chart(fig)


def plot_speed_analysis(df, selected_event, pre_event_range, post_event_range):
    df['VelKPH'] = pd.to_numeric(df['VelKPH'], errors='coerce')
    
    # Find the index of the selected event
    event_index = df[df['Event'] == selected_event].index[0]

    # Extract a window of rows around the event based on the pre-event and post-event range
    window_start = max(0, event_index - pre_event_range)
    window_end = min(df.shape[0], event_index + post_event_range)
    df_window = df.iloc[window_start:window_end]

    # Plot VelKPH values
    plt.figure(figsize=(12, 6))
    plt.plot(df_window['Distm'], df_window['VelKPH'], color='purple')
    plt.axvline(x=df_window.loc[event_index, 'Distm'], color='r', linestyle='--', label='Event')
    plt.title(f'Speed (VelKPH) around Event {selected_event}')
    plt.xlabel('Distm (Distance)')
    plt.ylabel('VelKPH')
    plt.grid(True)

    # Display the plot in Streamlit
    st.pyplot(plt.gcf())
    plt.close()

# We also need to modify the show_event_analysis_updated function to pass the offset to the plot_event_analysis_updated function
def show_event_analysis_with_scatter(df):
    """
    Display the analysis for selected event and row offset in the Streamlit app.
    """
    # Check if there are any events
    event_options = df[df['Event'].notnull()]['Event'].unique().tolist()
    if not event_options:
        st.write("No events found in the data.")
        return

    pre_event_range = st.sidebar.slider("Select Pre-Event Range", 100, 700, 100, 50)
    post_event_range = st.sidebar.slider("Select Post-Event Range", 100, 700, 100, 50)
    selected_event = st.sidebar.selectbox("Select an Event", event_options)

    # Display event description
    scenario = df['Scenario'].iloc[0]
    event_description = get_event_description(scenario, selected_event)
    st.title(event_description)  # Display as title
    
    # Allow users to select the row offset using a slider
    offset = st.sidebar.slider("Select Row Offset", -100, 100, 0)
    
    # Checkboxes to select which graphs to display
    show_wheeleang = st.sidebar.checkbox("Show WheeleAng Graph", True)
    show_thracce = st.sidebar.checkbox("Show ThrAcce Graph", True)
    show_brakacce = st.sidebar.checkbox("Show BrakAcce Graph", True)
    show_velkph = st.sidebar.checkbox("Show VelKPH Graph", True)
    
    # Plotting the selected graphs
    if show_wheeleang:
        plot_event_analysis_updated(df, selected_event, 'WheeleAng', offset, pre_event_range, post_event_range)
    if show_thracce:
        plot_event_analysis_updated(df, selected_event, 'ThrAcce', offset, pre_event_range, post_event_range)
    if show_brakacce:
        plot_event_analysis_updated(df, selected_event, 'BrakAcce', offset, pre_event_range, post_event_range)
    if show_velkph:
        plot_speed_analysis(df, selected_event, pre_event_range, post_event_range)

    matching_rows = df[df['Event'] == selected_event]
    event_row_index = matching_rows.index[0]
    changes = calculate_changes(df, event_row_index, offset)
    
    # Display the changes
    participant = df['Participant'].iloc[0]
    order = df['Order'].iloc[0]
    st.write(f"Participant {participant}_{order} changed the value of BrakAcce by {changes['BrakAcce']:.2f} points, ThrAcce by {changes['ThrAcce']:.2f} points, and WheeleAng by {changes['WheeleAng']:.2f} points.")
    st.write(f"The time difference is {changes['TimeDifference']} seconds and the distance difference is {changes['DistmDifference']} meters.")


# Streamlit app
def main():
    st.title("Driving Simulator Data Processor :car: :brain: :smile: \n by Eden Eldar")

    # Create a sidebar menu for navigation
    menu = ["Home", "Event Analysis"]
    choice = st.sidebar.selectbox("Menu", menu)

    # Common file upload for both Home and Event Analysis
    uploaded_file = st.file_uploader("Choose a file")

    if uploaded_file is not None:
        # Capture the original file name
        file_content = uploaded_file.getvalue().decode('utf-8')

        original_file_name = uploaded_file.name
        file_name = original_file_name.split(".")[0]

        # # Save the uploaded file to a temporary location
        # with open("temp.txt", "wb") as f:
        #     f.write(uploaded_file.getvalue())

        try:
            # Process the uploaded file
            df_sorted = process_raw_file_for_streamlit(file_content, original_file_name)

            if choice == "Home":
                # Display the processed data
                st.dataframe(df_sorted)
                st.subheader("Edit Event Highlight Values")
                scenario = df_sorted['Scenario'].iloc[0]
                current_values = HIGHLIGHT_VALUES.get(scenario, [])
                
                # Display a text box for each event
                modified_values = []
                for i, value in enumerate(current_values):
                    new_value = st.text_input(f"Event {i+1} Distm value", value=str(value))
                    try:
                        modified_values.append(float(new_value))
                    except ValueError:
                        st.error(f"Invalid input for Event {i+1}. Please enter a numeric value.")
                
                # Update HIGHLIGHT_VALUES when the button is pressed
                if st.button("Accept Changes"):
                    HIGHLIGHT_VALUES[scenario] = modified_values
                    df_sorted = process_raw_file_for_streamlit(file_content, original_file_name)
                    st.markdown(':green[The events distance values been update !')
                    # st.dataframe(df_sorted)

                scenario = df_sorted['Scenario'].iloc[0]

                # Save the processed data as an XLSX file with highlighting
                xlsx_path = save_as_xlsx_with_highlight_refined(df_sorted, scenario, file_name)

                # Offer option to download the sorted data
                if st.button("Download Sorted Data as XLSX"):
                    xlsx_path = save_as_xlsx_with_highlight_refined(df_sorted, scenario, file_name)
                    
                    with open(xlsx_path, "rb") as f:
                        b64 = base64.b64encode(f.read()).decode()  # Convert bytes to string
                        href = f'<a href="data:file/xlsx;base64,{b64}" download="{xlsx_path}">Download XLSX File</a>'
                        st.markdown(href, unsafe_allow_html=True)

            elif choice == "Event Analysis":
                show_event_analysis_with_scatter(df_sorted)

        except Exception as e:
            st.write("An error occurred:", str(e))
if __name__ == "__main__":
    main()
